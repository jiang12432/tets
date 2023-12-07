import openpyxl
from openpyxl.worksheet.worksheet import Worksheet


class ExcelHandler(object):
    """操作Excel"""
    def __init__(self, file):
        """初始化方法"""
        self.file = file

    def open_sheet(self, name) -> Worksheet:
        # 在函数后者方法的后面加一个箭头，表示此函数返回值是一个这样的类型
        """打开表单"""
        wb = openpyxl.load_workbook(self.file)
        sheet = wb[name]
        return sheet

    def header(self, sheet_name):
        """获取表头"""
        sheet = self.open_sheet(sheet_name)
        headers = []
        for i in sheet[1]:
            headers.append(i.value)
        return headers

    def read(self, _name):
        """读取所有的数据"""
        sheet = self.open_sheet(_name)
        rows = list(sheet.rows)
        data = []
        for row in rows[1:]:
            row_data = []
            for cell in row:
                row_data.append(cell.value)
                # 列表转字典
                data_dict = dict(zip(self.header(_name), row_data))
            data.append(data_dict)
        return data

    @staticmethod
    def write(file, sheet_name, row, column, data):
        """"写入Excel数据"""
        wb = openpyxl.load_workbook(file)
        sheet = wb[sheet_name]
        # 修改单元格
        sheet.cell(row, column).value = data
        # 保存
        wb.save(file)
        # 关闭
        wb.close()


if __name__ == "__main__":
    excle = ExcelHandler(r"C:\Users\86187\PycharmProjects\KD接口自动化\data\gama_data.xlsx")
    a = excle.read("Sheet1")
    print(a)
    # data = excle.read("Sheet1")
    # print(data)

# zip 打包
